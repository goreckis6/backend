#!/usr/bin/env python3
"""
CSV to XLSX Converter
Converts CSV files to Excel XLSX format using pandas and openpyxl.
"""

import pandas as pd
import argparse
import os
import sys
from datetime import datetime
import traceback

def create_xlsx_from_csv(csv_file, output_file, title="CSV Data", author="Unknown"):
    """
    Convert CSV file to XLSX format using pandas and openpyxl.
    
    Args:
        csv_file (str): Path to input CSV file
        output_file (str): Path to output XLSX file
        title (str): Document title
        author (str): Document author
    """
    print(f"Starting CSV to XLSX conversion...")
    print(f"Input: {csv_file}")
    print(f"Output: {output_file}")
    print(f"Title: {title}")
    print(f"Author: {author}")
    
    try:
        # Read CSV file
        print("Reading CSV file...")
        df = pd.read_csv(csv_file)
        print(f"CSV loaded successfully: {len(df)} rows, {len(df.columns)} columns")
        print(f"Column names: {list(df.columns)}")
        print(f"First few rows:")
        print(df.head())
        
        # Process all rows - no limits
        print(f"Processing all {len(df)} rows (including any repeated data)")
        
        # Check for any data issues
        print(f"Data types: {df.dtypes.to_dict()}")
        print(f"Any null values: {df.isnull().sum().sum()}")
        
        # Create Excel writer with XLSX format
        print("Creating XLSX file...")
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Write data to the first sheet
            df.to_excel(writer, sheet_name='Data', index=False)
            
            # Create a summary sheet
            summary_data = {
                'Metric': [
                    'Total Rows',
                    'Total Columns', 
                    'File Size (bytes)',
                    'Generated Date',
                    'Author',
                    'Title'
                ],
                'Value': [
                    len(df),
                    len(df.columns),
                    os.path.getsize(csv_file),
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    author,
                    title
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Create column info sheet
            column_info = []
            for i, col in enumerate(df.columns, 1):
                column_info.append({
                    'Column Number': i,
                    'Column Name': col,
                    'Data Type': str(df[col].dtype),
                    'Non-Null Count': df[col].count(),
                    'Null Count': df[col].isnull().sum(),
                    'Unique Values': df[col].nunique()
                })
            
            column_df = pd.DataFrame(column_info)
            column_df.to_excel(writer, sheet_name='Column Info', index=False)
            
            # Get the workbook and worksheet to add formatting
            workbook = writer.book
            data_sheet = writer.sheets['Data']
            summary_sheet = writer.sheets['Summary']
            column_sheet = writer.sheets['Column Info']
            
            # Auto-adjust column widths
            for sheet in [data_sheet, summary_sheet, column_sheet]:
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                    sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add header formatting to data sheet
            from openpyxl.styles import Font, PatternFill, Alignment
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in data_sheet[1]:  # First row (headers)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Add alternating row colors
            from openpyxl.styles import PatternFill
            light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            
            for row_num in range(2, len(df) + 2):  # Start from row 2 (skip header)
                if row_num % 2 == 0:  # Even rows
                    for cell in data_sheet[row_num]:
                        cell.fill = light_fill
        
        # Verify file was created
        if os.path.exists(output_file):
            file_size = os.path.getsize(output_file)
            print(f"XLSX file created successfully: {file_size} bytes")
            return True
        else:
            print("ERROR: XLSX file was not created")
            return False
            
    except Exception as e:
        print(f"ERROR: Failed to create XLSX from CSV: {e}")
        traceback.print_exc()
        return False

def main():
    parser = argparse.ArgumentParser(description='Convert CSV to XLSX format')
    parser.add_argument('csv_file', help='Input CSV file path')
    parser.add_argument('output_file', help='Output XLSX file path')
    parser.add_argument('--title', default='CSV Data', help='Document title')
    parser.add_argument('--author', default='Unknown', help='Document author')
    
    args = parser.parse_args()
    
    print("=== CSV to XLSX Converter ===")
    print(f"Python version: {sys.version}")
    print(f"Working directory: {os.getcwd()}")
    print(f"Arguments: {vars(args)}")
    
    # Check if input file exists
    if not os.path.exists(args.csv_file):
        print(f"ERROR: Input CSV file not found: {args.csv_file}")
        sys.exit(1)
    
    # Check pandas availability
    try:
        print(f"Pandas version: {pd.__version__}")
    except Exception as e:
        print(f"ERROR: Pandas not available: {e}")
        sys.exit(1)
    
    # Create output directory if it doesn't exist
    output_dir = os.path.dirname(args.output_file)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # Convert CSV to XLSX
    success = create_xlsx_from_csv(
        args.csv_file,
        args.output_file,
        args.title,
        args.author
    )
    
    if success:
        print("=== CONVERSION SUCCESSFUL ===")
        sys.exit(0)
    else:
        print("=== CONVERSION FAILED ===")
        sys.exit(1)

if __name__ == "__main__":
    main()
